"""
Excel File Splitter by Make
Uses Excel automation (win32com) for speed and full formatting preservation.
"""

import tkinter as tk
from tkinter import ttk, filedialog
from pathlib import Path
import threading
import queue
import gc
import shutil

import pandas as pd

# Try to import win32com for Excel automation
try:
    import win32com.client
    HAS_WIN32COM = True
except ImportError:
    HAS_WIN32COM = False

import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


class DarkTheme:
    BG = "#1e1e1e"
    BG_SECONDARY = "#2d2d2d"
    BG_TERTIARY = "#3c3c3c"
    FG = "#ffffff"
    FG_SECONDARY = "#b0b0b0"
    ACCENT = "#0078d4"
    SUCCESS = "#4caf50"
    ERROR = "#f44336"
    WARNING = "#ff9800"
    BORDER = "#404040"


class DropZone(tk.Canvas):
    def __init__(self, parent, on_drop_callback, **kwargs):
        super().__init__(parent, **kwargs)
        self.on_drop_callback = on_drop_callback
        self.configure(bg=DarkTheme.BG_SECONDARY, highlightthickness=2,
            highlightbackground=DarkTheme.BORDER, highlightcolor=DarkTheme.ACCENT)
        self._draw_drop_zone()
        self.bind("<Button-1>", self._on_click)
        self.bind("<Enter>", lambda e: self._draw_drop_zone(True))
        self.bind("<Leave>", lambda e: self._draw_drop_zone(False))
        
    def _draw_drop_zone(self, hover=False):
        self.delete("all")
        w, h = self.winfo_width() or 400, self.winfo_height() or 200
        color = DarkTheme.ACCENT if hover else DarkTheme.BORDER
        self.create_rectangle(10, 10, w-10, h-10, outline=color, width=2, dash=(10, 5))
        cx, cy = w // 2, h // 2 - 20
        icon_color = DarkTheme.ACCENT if hover else DarkTheme.FG_SECONDARY
        self.create_polygon(cx-25, cy-20, cx+15, cy-20, cx+25, cy-10, cx+25, cy+25, cx-25, cy+25, fill=icon_color)
        self.create_polygon(cx+15, cy-20, cx+15, cy-10, cx+25, cy-10, fill=DarkTheme.BG_SECONDARY, outline=icon_color)
        text_color = DarkTheme.FG if hover else DarkTheme.FG_SECONDARY
        self.create_text(cx, cy+50, text="Drop Excel file here\nor click to browse", fill=text_color, font=("Segoe UI", 12), justify="center")
        
    def _on_click(self, event):
        files = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if files:
            self.on_drop_callback(list(files))
            
    def update_size(self, event=None):
        self.after(10, self._draw_drop_zone)


class ExcelSplitterApp:
    INVALID_MAKES = {'%', 'aaa', 'hold', 'x', 'test', 'n/a', 'na', 'none', '', 'tbd', 'unknown', 'other', 'misc', 'temp', 'delete'}
    
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Splitter by Make")
        self.root.geometry("750x650")
        self.root.configure(bg=DarkTheme.BG)
        self.root.minsize(500, 400)
        self.message_queue = queue.Queue()
        self.processing = False
        self._build_ui()
        self._setup_drag_drop()
        self._process_queue()
        
        if HAS_WIN32COM:
            self.log("Excel automation available - fast mode enabled", "success")
        else:
            self.log("Install pywin32 for faster processing: pip install pywin32", "warning")
    
    def _setup_drag_drop(self):
        try:
            self.drop_zone.drop_target_register("DND_Files")
            self.drop_zone.dnd_bind("<<Drop>>", self._on_dnd_drop)
            self.log("Drag and drop enabled", "success")
        except Exception:
            self.log("Click to browse for files", "info")
    
    def _on_dnd_drop(self, event):
        files = self.root.tk.splitlist(event.data)
        if files:
            self._start_processing(list(files))
        
    def _build_ui(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Dark.TFrame", background=DarkTheme.BG)
        style.configure("Dark.TLabel", background=DarkTheme.BG, foreground=DarkTheme.FG, font=("Segoe UI", 10))
        style.configure("DarkTitle.TLabel", background=DarkTheme.BG, foreground=DarkTheme.FG, font=("Segoe UI", 16, "bold"))
        style.configure("DarkSubtitle.TLabel", background=DarkTheme.BG, foreground=DarkTheme.FG_SECONDARY, font=("Segoe UI", 9))
        style.configure("TProgressbar", background=DarkTheme.ACCENT, troughcolor=DarkTheme.BG_SECONDARY, thickness=10)
        
        main_frame = ttk.Frame(self.root, style="Dark.TFrame", padding=20)
        main_frame.pack(fill="both", expand=True)
        
        ttk.Label(main_frame, text="Excel Splitter by Make", style="DarkTitle.TLabel").pack(pady=(0, 5))
        ttk.Label(main_frame, text="Uses Excel automation - preserves ALL formatting", style="DarkSubtitle.TLabel").pack(pady=(0, 20))
        
        drop_frame = ttk.Frame(main_frame, style="Dark.TFrame")
        drop_frame.pack(fill="both", expand=True, pady=(0, 20))
        self.drop_zone = DropZone(drop_frame, self._start_processing, width=400, height=150)
        self.drop_zone.pack(fill="both", expand=True)
        self.drop_zone.bind("<Configure>", self.drop_zone.update_size)
        
        output_frame = ttk.Frame(main_frame, style="Dark.TFrame")
        output_frame.pack(fill="x", pady=(0, 10))
        ttk.Label(output_frame, text="Output:", style="Dark.TLabel").pack(side="left")
        self.output_var = tk.StringVar(value=str(Path.home() / "Desktop"))
        tk.Entry(output_frame, textvariable=self.output_var, bg=DarkTheme.BG_SECONDARY, fg=DarkTheme.FG,
            insertbackground=DarkTheme.FG, relief="flat", font=("Segoe UI", 10)).pack(side="left", fill="x", expand=True, padx=(10, 10))
        tk.Button(output_frame, text="Browse", command=lambda: self.output_var.set(filedialog.askdirectory(initialdir=self.output_var.get()) or self.output_var.get()),
            bg=DarkTheme.BG_TERTIARY, fg=DarkTheme.FG, relief="flat", font=("Segoe UI", 9)).pack(side="right")
        
        # Version selection frame with two dropdowns
        version_frame = ttk.Frame(main_frame, style="Dark.TFrame")
        version_frame.pack(fill="x", pady=(0, 15))
        ttk.Label(version_frame, text="Version:  v", style="Dark.TLabel").pack(side="left")
        
        # Major version dropdown (before decimal)
        self.version_major_var = tk.StringVar(value="1")
        major_options = [str(i) for i in range(1, 10)]
        self.version_major_dropdown = ttk.Combobox(version_frame, textvariable=self.version_major_var, 
            values=major_options, width=3, state="readonly", font=("Segoe UI", 10))
        self.version_major_dropdown.pack(side="left")
        
        ttk.Label(version_frame, text=".", style="Dark.TLabel").pack(side="left")
        
        # Minor version dropdown (after decimal)
        self.version_minor_var = tk.StringVar(value="0")
        minor_options = [str(i) for i in range(0, 10)]
        self.version_minor_dropdown = ttk.Combobox(version_frame, textvariable=self.version_minor_var, 
            values=minor_options, width=3, state="readonly", font=("Segoe UI", 10))
        self.version_minor_dropdown.pack(side="left")
        
        # Style the comboboxes
        self.root.option_add('*TCombobox*Listbox.background', DarkTheme.BG_SECONDARY)
        self.root.option_add('*TCombobox*Listbox.foreground', DarkTheme.FG)
        
        self.progress_var = tk.DoubleVar()
        ttk.Progressbar(main_frame, variable=self.progress_var, mode="determinate").pack(fill="x", pady=(0, 10))
        
        # Status and open folder button frame
        status_frame = ttk.Frame(main_frame, style="Dark.TFrame")
        status_frame.pack(fill="x", pady=(0, 15))
        
        self.status_var = tk.StringVar(value="Ready - Drop master Excel file")
        ttk.Label(status_frame, textvariable=self.status_var, style="DarkSubtitle.TLabel").pack(side="left")
        
        # Open folder button (hidden initially)
        self.open_folder_btn = tk.Button(status_frame, text="Open Folder", command=self._open_output_folder,
            bg=DarkTheme.ACCENT, fg=DarkTheme.FG, relief="flat", font=("Segoe UI", 9), cursor="hand2")
        self.output_folder_path = None  # Will be set when processing completes
        
        log_frame = ttk.Frame(main_frame, style="Dark.TFrame")
        log_frame.pack(fill="both", expand=True)
        ttk.Label(log_frame, text="Log:", style="Dark.TLabel").pack(anchor="w")
        
        log_container = ttk.Frame(log_frame, style="Dark.TFrame")
        log_container.pack(fill="both", expand=True, pady=(5, 0))
        scrollbar = tk.Scrollbar(log_container)
        scrollbar.pack(side="right", fill="y")
        self.log_text = tk.Text(log_container, height=12, bg=DarkTheme.BG_SECONDARY, fg=DarkTheme.FG,
            relief="flat", font=("Consolas", 9), wrap="word", yscrollcommand=scrollbar.set)
        self.log_text.pack(fill="both", expand=True)
        scrollbar.config(command=self.log_text.yview)
        
        for tag, color in [("success", DarkTheme.SUCCESS), ("error", DarkTheme.ERROR), ("warning", DarkTheme.WARNING), ("info", DarkTheme.ACCENT)]:
            self.log_text.tag_configure(tag, foreground=color)
        
    def _process_queue(self):
        try:
            while True:
                msg_type, data = self.message_queue.get_nowait()
                if msg_type == "log":
                    self.log_text.insert("end", data[0] + "\n", data[1])
                    self.log_text.see("end")
                elif msg_type == "progress":
                    self.progress_var.set(data)
                elif msg_type == "status":
                    self.status_var.set(data)
                elif msg_type == "done":
                    self.processing = False
                elif msg_type == "show_button":
                    self.open_folder_btn.pack(side="right", padx=(10, 0))
        except queue.Empty:
            pass
        self.root.after(50, self._process_queue)
            
    def log(self, msg, tag=None):
        self.message_queue.put(("log", (msg, tag)))
    
    def _open_output_folder(self):
        if self.output_folder_path and self.output_folder_path.exists():
            import subprocess
            subprocess.Popen(f'explorer "{self.output_folder_path}"')
    
    def _show_open_folder_button(self, folder_path):
        self.output_folder_path = folder_path
        self.message_queue.put(("show_button", None))
        
    def _is_valid_make(self, make):
        if make is None:
            return False
        m = str(make).strip().lower()
        return m and m not in self.INVALID_MAKES and len(m) > 1

    def _start_processing(self, file_paths):
        if self.processing:
            return
        self.processing = True
        self.log_text.delete("1.0", "end")
        self.progress_var.set(0)
        version = f"v{self.version_major_var.get()}.{self.version_minor_var.get()}"
        # Hide open folder button when starting new process
        self.open_folder_btn.pack_forget()
        threading.Thread(target=self._process_file, args=(file_paths[0], self.output_var.get(), version), daemon=True).start()
        
    def _process_file(self, file_path, output_dir, version):
        try:
            file_path = Path(file_path)
            base_output_dir = Path(output_dir)
            
            # Remove 'v' prefix if present for version number
            version_num = version.replace('v', '')
            
            # Parse source filename to find base name (everything before the version)
            # e.g., "1 - All ID³ Manufacturer Map v1.0.0.0" -> "1 - All ID³ Manufacturer Map"
            import re
            stem = file_path.stem
            # Find everything before " v" followed by numbers/dots at the end
            match = re.match(r'^(.+?)\s*v[\d.]+$', stem, re.IGNORECASE)
            if match:
                base_name = match.group(1).strip()
            else:
                # No version found, use full stem
                base_name = stem
            
            # Create folder with base name + selected version
            # e.g., "1 - All ID³ Manufacturer Map v1.1"
            folder_name = f"{base_name} {version}"
            output_dir = base_output_dir / folder_name
            output_dir.mkdir(parents=True, exist_ok=True)
            
            # Store version for file naming
            self.current_version = version_num
            
            self.log(f"Source: {file_path.name}", "info")
            self.log(f"Output: {folder_name}\n", "info")
            
            # Step 1: Scan for makes using pandas (fast)
            self.log("Scanning for Makes...", "info")
            self.message_queue.put(("status", "Scanning for Makes..."))
            
            excel = pd.ExcelFile(file_path)
            all_makes = set()
            skipped = set()
            sheet_make_cols = {}  # sheet_name -> (make_column_index, make_column_letter)
            
            for sheet in excel.sheet_names:
                df = pd.read_excel(excel, sheet_name=sheet, nrows=0)
                make_col_idx = None
                for idx, col in enumerate(df.columns):
                    if str(col).strip().lower() == 'make':
                        make_col_idx = idx + 1  # 1-based
                        break
                
                if make_col_idx:
                    # Convert to Excel column letter
                    col_letter = ""
                    temp = make_col_idx
                    while temp > 0:
                        temp, remainder = divmod(temp - 1, 26)
                        col_letter = chr(65 + remainder) + col_letter
                    
                    sheet_make_cols[sheet] = (make_col_idx, col_letter)
                    df_full = pd.read_excel(excel, sheet_name=sheet, usecols=[make_col_idx - 1])
                    for val in df_full.iloc[:, 0].dropna().unique():
                        v = str(val).strip()
                        if self._is_valid_make(v):
                            all_makes.add(v)
                        elif v:
                            skipped.add(v)
            
            excel.close()
            
            if not all_makes:
                self.log("No valid Makes found!", "error")
                return
            
            self.log(f"Found {len(all_makes)} Makes", "info")
            if skipped:
                self.log(f"Skipped invalid: {', '.join(sorted(skipped))}", "warning")
            self.log(f"Sheets with Make column: {len(sheet_make_cols)}\n", "info")
            
            if HAS_WIN32COM:
                try:
                    self._process_with_excel(file_path, output_dir, all_makes, sheet_make_cols)
                except Exception as e:
                    self.log(f"Excel automation failed: {e}", "error")
                    self.log("Falling back to openpyxl...", "warning")
                    self._process_with_openpyxl(file_path, output_dir, all_makes, sheet_make_cols)
            else:
                self.log("win32com not available, using openpyxl", "warning")
                self._process_with_openpyxl(file_path, output_dir, all_makes, sheet_make_cols)
            
        except Exception as e:
            self.log(f"Error: {e}", "error")
            import traceback
            self.log(traceback.format_exc(), "error")
        finally:
            self.message_queue.put(("done", None))
    
    def _process_with_excel(self, file_path, output_dir, all_makes, sheet_make_cols):
        """Process using Excel automation - FAST and preserves ALL formatting"""
        self.log("Using Excel automation (fast mode)...\n", "info")
        
        excel_app = None
        try:
            # Use Dispatch instead of gencache for better reliability
            excel_app = win32com.client.Dispatch('Excel.Application')
            excel_app.Visible = False
            excel_app.DisplayAlerts = False
            excel_app.ScreenUpdating = False
            self.log("Excel connected successfully", "success")
            
            total = len(all_makes)
            for idx, make in enumerate(sorted(all_makes)):
                output_file = output_dir / f"{make} ID³ Manufacturer Map v{self.current_version}.xlsx"
                self.log(f"[{idx+1}/{total}] {make}...", None)
                self.message_queue.put(("status", f"Processing {make} ({idx+1}/{total})..."))
                
                try:
                    # Copy source file
                    shutil.copy2(file_path, output_file)
                    
                    # Open in Excel
                    wb = excel_app.Workbooks.Open(str(output_file.absolute()))
                    
                    for sheet_name, (make_col_idx, col_letter) in sheet_make_cols.items():
                        try:
                            ws = wb.Sheets(sheet_name)
                            last_row = ws.Cells(ws.Rows.Count, make_col_idx).End(-4162).Row  # xlUp = -4162
                            
                            if last_row <= 1:
                                continue
                            
                            # Check if AutoFilter already exists, turn it off temporarily
                            had_autofilter = ws.AutoFilterMode
                            if had_autofilter:
                                ws.AutoFilterMode = False
                            
                            # Use AutoFilter to bulk delete non-matching rows (FAST)
                            data_range = ws.Range(ws.Cells(1, 1), ws.Cells(last_row, ws.UsedRange.Columns.Count))
                            
                            # Apply AutoFilter on Make column, filter for NOT equal to target make
                            data_range.AutoFilter(Field=make_col_idx, Criteria1="<>" + make)
                            
                            # Delete all visible (filtered) rows except header
                            try:
                                visible_range = ws.Range(ws.Cells(2, 1), ws.Cells(last_row, 1)).SpecialCells(12)  # xlCellTypeVisible = 12
                                visible_range.EntireRow.Delete()
                            except:
                                pass  # No rows to delete (all match)
                            
                            # Show all data (clears filter but keeps AutoFilter enabled)
                            try:
                                ws.ShowAllData()
                            except:
                                pass  # No filter active
                            
                        except Exception as e:
                            pass
                    
                    wb.Save()
                    wb.Close(False)
                    self.log(f" Done", "success")
                    
                except Exception as e:
                    self.log(f" Error: {e}", "error")
                    try:
                        wb.Close(False)
                    except:
                        pass
                
                self.message_queue.put(("progress", ((idx + 1) / total) * 100))
            
            self.log(f"\nComplete! Created {total} files.", "success")
            self.message_queue.put(("status", "Complete!"))
            self._show_open_folder_button(output_dir)
            
        finally:
            if excel_app:
                try:
                    excel_app.ScreenUpdating = True
                    excel_app.DisplayAlerts = True
                    excel_app.Quit()
                except:
                    pass
    
    def _process_with_openpyxl(self, file_path, output_dir, all_makes, sheet_make_cols):
        """Fallback: Process using openpyxl (slower)"""
        self.log("Using openpyxl (slower mode)...\n", "warning")
        
        from openpyxl import load_workbook
        
        total = len(all_makes)
        for idx, make in enumerate(sorted(all_makes)):
            output_file = output_dir / f"{make} ID³ Manufacturer Map v{self.current_version}.xlsx"
            self.log(f"[{idx+1}/{total}] {make}...", None)
            self.message_queue.put(("status", f"Processing {make} ({idx+1}/{total})..."))
            
            try:
                shutil.copy2(file_path, output_file)
                wb = load_workbook(output_file)
                
                for sheet_name, (make_col_idx, _) in sheet_make_cols.items():
                    if sheet_name not in wb.sheetnames:
                        continue
                    ws = wb[sheet_name]
                    
                    rows_to_delete = []
                    for row in range(ws.max_row, 1, -1):
                        cell_val = ws.cell(row=row, column=make_col_idx).value
                        if cell_val is None or str(cell_val).strip().lower() != make.lower():
                            rows_to_delete.append(row)
                    
                    for row in rows_to_delete:
                        ws.delete_rows(row)
                
                wb.save(output_file)
                wb.close()
                self.log(f" Done", "success")
                
            except Exception as e:
                self.log(f" Error: {e}", "error")
            
            self.message_queue.put(("progress", ((idx + 1) / total) * 100))
            gc.collect()
        
        self.log(f"\nComplete! Created {total} files.", "success")
        self.message_queue.put(("status", "Complete!"))
        self._show_open_folder_button(output_dir)


def main():
    try:
        from tkinterdnd2 import TkinterDnD
        root = TkinterDnD.Tk()
    except ImportError:
        root = tk.Tk()
    ExcelSplitterApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
